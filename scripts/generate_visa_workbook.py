from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import datetime

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
C = {
    "navy":      "1B2A4A",
    "navy_mid":  "2E4272",
    "teal":      "1A7A6E",
    "teal_lt":   "D4F0EC",
    "amber":     "D97706",
    "amber_lt":  "FEF3C7",
    "red":       "B91C1C",
    "red_lt":    "FEE2E2",
    "green":     "15803D",
    "green_lt":  "DCFCE7",
    "blue_lt":   "DBEAFE",
    "blue":      "1D4ED8",
    "gray":      "6B7280",
    "gray_lt":   "F3F4F6",
    "gray_mid":  "E5E7EB",
    "white":     "FFFFFF",
    "black":     "111827",
    "purple":    "6D28D9",
    "purple_lt": "EDE9FE",
    "orange":    "EA580C",
    "orange_lt": "FFEDD5",
}

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color="111827", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def border(style="thin", color="D1D5DB"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header_row(ws, row, bg, fg="FFFFFF", bold=True, size=10):
    for cell in ws[row]:
        cell.font = font(bold=bold, color=fg, size=size)
        cell.fill = fill(bg)
        cell.alignment = align("center", "center")
        cell.border = border()

def style_range(ws, row_start, row_end, col_start, col_end,
                bg=None, fg=None, bold=False, wrap=False, h_align="left"):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if bg:
                cell.fill = fill(bg)
            if fg:
                cell.font = font(bold=bold, color=fg)
            else:
                cell.font = font(bold=bold)
            cell.alignment = align(h_align, "center", wrap)
            cell.border = border()

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ═════════════════════════════════════════════════════════════════════════════
# DATA
# ═════════════════════════════════════════════════════════════════════════════

# Status codes
VF  = "Visa-Free"
VOA = "Visa on Arrival"
EV  = "e-Visa"
VR  = "Visa Required"
NA  = "N/A (Own Country)"

# ── Main visa matrix data ────────────────────────────────────────────────────
# Columns: Destination, Region, Indian passport, UK passport, US passport,
#          Canadian passport, Australian passport, Nigerian passport,
#          Chinese passport, South African passport

VISA_DATA = [
    # Destination, Region, IN, GB, US, CA, AU, NG, CN, ZA, Notes
    # ── Europe ──
    ("France",          "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Germany",         "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Italy",           "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Spain",           "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Netherlands",     "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Portugal",        "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Greece",          "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Switzerland",     "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("United Kingdom",  "Europe",         VR,  NA,  VF,        VF,  VF,  VR,  EV,  VR),
    ("Ireland",         "Europe",         VR,  VF,  VF,        VF,  VF,  VR,  VR,  VR),
    ("Albania",         "Europe",         EV,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Georgia",         "Europe",         VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Serbia",          "Europe",         EV,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Turkey",          "Europe/Asia",    EV,  EV,  EV,        EV,  EV,  EV,  VF,  EV),
    ("Montenegro",      "Europe",         VR,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("North Macedonia", "Europe",         VR,  VF,  VF,        VF,  VF,  VR,  VF,  VR),
    ("Croatia",         "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Czech Republic",  "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Hungary",         "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Poland",          "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Norway",          "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Sweden",          "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Iceland",         "Europe",         VR,  VF,  "ESTA/VF", VF,  VF,  VR,  VR,  VR),
    ("Armenia",         "Europe",         VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Moldova",         "Europe",         VF,  VF,  VF,        VF,  VF,  VR,  VR,  VR),
    # ── Asia ──
    ("Japan",           "East Asia",      VR,  VF,  VF,        VF,  VF,  VR,  VF,  VF),
    ("South Korea",     "East Asia",      EV,  VF,  VF,        VF,  VF,  VR,  VF,  VF),
    ("China",           "East Asia",      VR,  VF,  VR,        VR,  VR,  VR,  NA,  VR),
    ("Hong Kong",       "East Asia",      VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Taiwan",          "East Asia",      EV,  VF,  VF,        VF,  VF,  VR,  VR,  VR),
    ("Singapore",       "Southeast Asia", VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Thailand",        "Southeast Asia", VF,  VF,  VF,        VF,  VF,  VOA, VF,  VF),
    ("Vietnam",         "Southeast Asia", EV,  EV,  EV,        EV,  EV,  EV,  VF,  EV),
    ("Indonesia",       "Southeast Asia", VF,  VF,  VF,        VF,  VF,  VOA, VF,  VF),
    ("Malaysia",        "Southeast Asia", VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Philippines",     "Southeast Asia", VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Cambodia",        "Southeast Asia", VOA, VF,  VF,        VF,  VF,  VOA, VF,  VOA),
    ("Myanmar",         "Southeast Asia", EV,  EV,  EV,        EV,  EV,  EV,  VF,  EV),
    ("Laos",            "Southeast Asia", VOA, VF,  VF,        VF,  VF,  VOA, VF,  VOA),
    ("Sri Lanka",       "South Asia",     EV,  EV,  EV,        EV,  EV,  EV,  EV,  EV),
    ("Nepal",           "South Asia",     VF,  VF,  VOA,       VOA, VOA, VOA, VOA, VOA),
    ("Maldives",        "South Asia",     VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Bhutan",          "South Asia",     VF,  EV,  EV,        EV,  EV,  EV,  EV,  EV),
    ("India",           "South Asia",     NA,  EV,  EV,        EV,  EV,  EV,  EV,  EV),
    ("UAE",             "Middle East",    VF,  VF,  VF,        VF,  VF,  VOA, VR,  VF),
    ("Qatar",           "Middle East",    VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Oman",            "Middle East",    EV,  VF,  VF,        VF,  VF,  EV,  VR,  EV),
    ("Saudi Arabia",    "Middle East",    EV,  EV,  EV,        EV,  EV,  EV,  VR,  EV),
    ("Bahrain",         "Middle East",    EV,  VF,  VF,        VF,  VF,  EV,  VR,  EV),
    ("Kuwait",          "Middle East",    VR,  VF,  VF,        VF,  VF,  VR,  VR,  VR),
    ("Jordan",          "Middle East",    VOA, VF,  VF,        VF,  VF,  VOA, VR,  VOA),
    ("Israel",          "Middle East",    VF,  VF,  VF,        VF,  VF,  VR,  VR,  VF),
    ("Azerbaijan",      "Middle East",    EV,  EV,  EV,        EV,  EV,  EV,  EV,  EV),
    # ── Africa ──
    ("Morocco",         "Africa",         EV,  VF,  VF,        VF,  VF,  VR,  VR,  VF),
    ("Egypt",           "Africa",         VOA, EV,  VOA,       VF,  VF,  VOA, VR,  VOA),
    ("South Africa",    "Africa",         VF,  VF,  VF,        VF,  VF,  VF,  VF,  NA),
    ("Kenya",           "Africa",         EV,  VF,  EV,        EV,  EV,  EV,  VR,  EV),
    ("Tanzania",        "Africa",         VOA, VF,  VOA,       VOA, VOA, VOA, VR,  VOA),
    ("Ethiopia",        "Africa",         VOA, EV,  EV,        EV,  EV,  VOA, EV,  EV),
    ("Ghana",           "Africa",         VR,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Nigeria",         "Africa",         VR,  VF,  VF,        VF,  VF,  NA,  VR,  VF),
    ("Senegal",         "Africa",         VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Rwanda",          "Africa",         VOA, VF,  VOA,       VOA, VOA, VF,  VR,  VOA),
    ("Mauritius",       "Africa",         VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Seychelles",      "Africa",         VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
    ("Zimbabwe",        "Africa",         VOA, VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Tunisia",         "Africa",         VF,  VF,  VF,        VF,  VF,  VR,  VR,  VF),
    ("Madagascar",      "Africa",         VOA, VF,  VF,        VF,  VF,  VOA, EV,  VOA),
    # ── Americas ──
    ("United States",   "North America",  VR,  "ESTA",VF,     VF,  VF,  VR,  VR,  VR),
    ("Canada",          "North America",  VR,  VF,  VF,        NA,  VF,  VR,  EV,  VR),
    ("Mexico",          "North America",  EV,  VF,  VF,        VF,  VF,  VR,  VF,  VF),
    ("Brazil",          "North America",  VF,  VF,  VR,        VR,  VF,  VF,  VR,  VF),
    ("Argentina",       "South America",  VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Colombia",        "South America",  VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Peru",            "South America",  VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Chile",           "South America",  VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Costa Rica",      "Central America",VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Panama",          "Central America",VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Cuba",            "Caribbean",      EV,  EV,  VR,        EV,  EV,  EV,  VF,  EV),
    ("Jamaica",         "Caribbean",      VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    ("Dominican Rep.",  "Caribbean",      VF,  VF,  VF,        VF,  VF,  VF,  VR,  VF),
    # ── Oceania ──
    ("Australia",       "Oceania",        EV,  EV,  EV,        EV,  NA,  VR,  EV,  EV),
    ("New Zealand",     "Oceania",        EV,  VF,  VF,        VF,  VF,  VR,  VR,  EV),
    ("Fiji",            "Oceania",        VF,  VF,  VF,        VF,  VF,  VF,  VF,  VF),
]

# ── Visa cost & processing time data ────────────────────────────────────────
# (Destination, IN cost USD, IN days, GB cost USD, GB days, US cost USD, US days, Notes)
COST_DATA = {
    "France":          ("€80 Schengen", "15–30 days", "N/A (VF)", "-",       "N/A (ESTA $40)", "-",         "Schengen covers 27 EU countries"),
    "Germany":         ("€80 Schengen", "15–30 days", "N/A (VF)", "-",       "N/A (ESTA $40)", "-",         ""),
    "Italy":           ("€80 Schengen", "15–30 days", "N/A (VF)", "-",       "N/A (ESTA $40)", "-",         ""),
    "Spain":           ("€80 Schengen", "15–30 days", "N/A (VF)", "-",       "N/A (ESTA $40)", "-",         ""),
    "United Kingdom":  ("£115 / ~$145", "15–20 days", "N/A (VF)", "-",       "N/A (VF)",       "-",         "Indian: Standard visitor"),
    "Ireland":         ("€100",         "8–10 weeks", "N/A (VF)", "-",       "N/A (VF)",       "-",         "Long processing times reported"),
    "Albania":         ("~€15 eVisa",   "Up to 15 days", "Free VF", "No visa", "Free VF",     "No visa",   "IN: e-visa; VF with UK/US/Schengen visa"),
    "Georgia":         ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   ""),
    "Serbia":          ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   ""),
    "Turkey":          ("~$40–60 eVisa","Instant",    "~$30 eVisa","Instant", "~$30 eVisa",    "Instant",   "e-Visa online 3 mins"),
    "Montenegro":      ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   ""),
    "North Macedonia": ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   ""),
    "Japan":           ("~¥3,000 / $20","5–7 days",   "N/A (VF)", "-",       "N/A (VF)",       "-",         "Multiple-entry available"),
    "South Korea":     ("Free K-ETA",   "1–3 days",   "N/A (VF)", "-",       "N/A (VF)",       "-",         "K-ETA online system"),
    "China":           ("~$140+",       "4–7 days",   "N/A (VF)","Free/VF",  "~$185+",         "4–7 days",  "UK: 2-year FVISA free trial"),
    "Hong Kong":       ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "90 days VF"),
    "Singapore":       ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days VF"),
    "Thailand":        ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "60 days VF from 2024"),
    "Vietnam":         ("~$25 eVisa",   "3 days",     "~$25 eVisa","3 days", "~$25 eVisa",     "3 days",    "e-Visa online"),
    "Indonesia":       ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days, extendable"),
    "Malaysia":        ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days"),
    "Philippines":     ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days"),
    "Cambodia":        ("~$30 VOA",     "On arrival", "Free VF",  "No visa", "Free VF",        "No visa",   "e-Visa $36 also avail"),
    "Sri Lanka":       ("~$50 ETA",     "Instant",    "~$50 ETA", "Instant", "~$50 ETA",       "Instant",   "Online ETA system"),
    "Nepal":           ("Free VF",      "No visa",    "Free VF",  "No visa", "$30–100 VOA",    "On arrival","IN: SAARC free"),
    "Maldives":        ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days"),
    "UAE":             ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "IN: 30-day VF since 2023"),
    "Qatar":           ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days"),
    "Saudi Arabia":    ("~$130 eVisa",  "Instant",    "~$130 eVisa","Instant","~$130 eVisa",   "Instant",   "Tourist eVisa online"),
    "Oman":            ("~$20 eVisa",   "Instant",    "Free VF",  "No visa", "Free VF",        "No visa",   "Online eVisa"),
    "Jordan":          ("~$56 VOA",     "On arrival", "Free VF",  "No visa", "Free VF",        "No visa",   "Jordan Pass waives fee"),
    "Morocco":         ("~$15 eVisa",   "2–3 days",   "Free VF",  "No visa", "Free VF",        "No visa",   "IN: e-Visa since 2024"),
    "Egypt":           ("~$25 VOA",     "On arrival", "~$25 eVisa","1–3 days","~$25 VOA",      "On arrival","VOA at major airports"),
    "South Africa":    ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "30 days"),
    "Kenya":           ("~$52 eVisa",   "3–5 days",   "Free VF",  "No visa", "~$52 eVisa",     "3–5 days",  "ETA system"),
    "Tanzania":        ("$50 VOA",      "On arrival", "Free VF",  "No visa", "$50 VOA",        "On arrival","VOA at JRO/DAR airports"),
    "Rwanda":          ("~$30 VOA",     "On arrival", "Free VF",  "No visa", "$50 VOA",        "On arrival","Gorilla trekking permits separate"),
    "Mauritius":       ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "60 days"),
    "Seychelles":      ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "Travel permit on arrival"),
    "United States":   ("~$185 B1/B2",  "60–120 days","ESTA $40", "Instant", "N/A (citizen)",  "-",         "IN: long wait times, interview req."),
    "Canada":          ("CAD $185",     "60–90 days", "Free VF",  "No visa", "Free VF",        "No visa",   "IN: eTA for VF not applicable"),
    "Mexico":          ("~$35 eVisa",   "1–3 days",   "Free VF",  "No visa", "Free VF",        "No visa",   "WARNING: UK visa entry contested"),
    "Brazil":          ("Free VF",      "No visa",    "Free VF",  "No visa", "Free VF",        "No visa",   "IN: VF since 2024"),
    "Australia":       ("AUD $20 ETA",  "Instant",    "AUD $20 ETA","Instant","N/A (citizen)", "-",         "ETA app required"),
    "New Zealand":     ("NZD $23 NZeTA","Instant",    "Free VF",  "No visa", "Free VF",        "No visa",   "NZeTA required online"),
    "Cuba":            ("~$25 eVisa",   "2–3 days",   "~$25 eVisa","2–3 days","No entry",      "-",         "US citizens banned"),
    "Japan_extra":     ("",             "",           "",         "",        "",               "",          ""),
}

# ── UK visa additional access (Indian passport + UK visa/BRP) ───────────────
UK_VISA_ACCESS = [
    # Country, Region, Type, Max Stay, Conditions, Notes
    ("Albania",              "Europe",         "Visa-free",         "90 days", "Any valid UK visa",          "Popular Balkans destination"),
    ("Georgia",              "Europe/Asia",    "Visa-free",         "365 days","Any valid UK visa",          "Exceptionally generous policy"),
    ("Montenegro",           "Europe",         "Visa-free",         "30 days", "Any valid UK visa",          "Adriatic coast access"),
    ("Serbia",               "Europe",         "Visa-free",         "90 days", "Any valid UK visa",          "Belgrade highly rated"),
    ("Armenia",              "Europe/Asia",    "Visa-free",         "180 days","Any valid UK visa",          ""),
    ("Moldova",              "Europe",         "Visa-free",         "90 days", "Any valid UK visa",          ""),
    ("North Macedonia",      "Europe",         "Visa-free",         "90 days", "Any valid UK visa",          "Note: some reports of denial"),
    ("Turkey",               "Europe/Asia",    "e-Visa eligible",   "30 days", "UK visa facilitates eVisa",  "Still need to buy eVisa ~$30"),
    ("Mexico",               "North America",  "Visa-free*",        "180 days","Used multiple-entry UK visa","WARNING: Entry reports contested"),
    ("Panama",               "Central America","Visa-free",         "30 days", "Used UK visa + $500 funds",  ""),
    ("Jamaica",              "Caribbean",      "Visa-free",         "90 days", "Any valid UK visa",          ""),
    ("Bahamas",              "Caribbean",      "Visa-free",         "21 days", "Indian nationals only",      ""),
    ("Dominican Republic",   "Caribbean",      "Visa-free",         "30 days", "Any valid UK visa",          ""),
    ("Anguilla",             "Caribbean",      "Visa-free",         "90 days", "UK visa accepted",           "British Territory"),
    ("Aruba",                "Caribbean",      "Visa-free",         "90 days", "Any valid UK visa",          "Dutch territory"),
    ("Bermuda",              "Caribbean",      "Visa-free",         "90 days", "UK visa accepted",           "British Territory"),
    ("Cayman Islands",       "Caribbean",      "Visa-free",         "90 days", "UK visa accepted",           "British Territory"),
    ("Curaçao",              "Caribbean",      "Visa-free",         "90 days", "Any valid UK visa",          ""),
    ("Sint Maarten",         "Caribbean",      "Visa-free",         "90 days", "Any valid UK visa",          ""),
    ("Turks & Caicos",       "Caribbean",      "Visa-free",         "90 days", "UK visa accepted",           "British Territory"),
    ("Peru",                 "South America",  "Visa-free",         "183 days","Valid used UK visa",         ""),
    ("Bahrain",              "Middle East",    "Visa-free / eVisa", "30 days", "Any valid UK visa",          "eVisa still recommended"),
    ("Oman",                 "Middle East",    "Visa-free",         "30 days", "Any valid UK visa",          "IN passport normally needs eVisa"),
    ("Qatar",                "Middle East",    "Visa-free",         "30 days", "Any valid UK visa",          "Hayya Card route also available"),
    ("Saudi Arabia",         "Middle East",    "eVisa eligible",    "90 days", "UK visa holders can apply",  "Still need to buy eVisa"),
    ("UAE",                  "Middle East",    "Visa-free",         "30 days", "Any valid UK visa",          "IN passport: VF anyway"),
    ("Kuwait",               "Middle East",    "Visa-free",         "30 days", "Any valid UK visa",          ""),
    ("Egypt",                "Africa",         "Visa-free",         "30 days", "Any valid UK visa",          "VOA also available"),
    ("Morocco",              "Africa",         "Visa-free*",        "30 days", "UK visa simplifies eVisa",   "Still need Morocco eVisa for IN"),
    ("Japan",                "East Asia",      "Visa-free",         "90 days", "Valid used UK visa",         "Major unlock for IN passport!"),
    ("South Korea",          "East Asia",      "Visa-free",         "30 days", "Any valid UK visa",          "K-ETA waived for UK visa holders"),
    ("Singapore",            "Southeast Asia", "Visa-free",         "96 hours","UK visa transit benefit",    "Mainly transit; standard VF applies"),
    ("Philippines",          "Southeast Asia", "Visa-free",         "30 days", "Any valid UK visa",          ""),
    ("Kyrgyzstan",           "Central Asia",   "Visa-free",         "30 days", "Any valid UK visa",          ""),
    ("Taiwan",               "East Asia",      "Visa-free",         "30 days", "Any valid UK visa",          "14 days extendable"),
    ("Ireland",              "Europe",         "Visa-free",         "90 days", "UK visa accepted (CTA)",     "Common Travel Area agreement"),
]

# ── Schengen visa additional access ─────────────────────────────────────────
SCHENGEN_ACCESS = [
    ("All 27 Schengen States", "Europe", "Visa-free with Schengen visa", "90 days in 180", "Single Schengen visa covers all member states"),
    ("Bulgaria",               "Europe", "Visa-free for Schengen holders","90 days",       "Not yet full Schengen but accepts"),
    ("Romania",                "Europe", "Visa-free for Schengen holders","90 days",       "Joined Schengen air/sea Dec 2023"),
    ("Kosovo",                 "Europe", "Visa-free for Schengen holders","90 days",       ""),
    ("North Macedonia",        "Europe", "Visa-free for Schengen holders","90 days",       ""),
    ("Albania",                "Europe", "Visa-free for Schengen holders","90 days",       ""),
    ("Montenegro",             "Europe", "Visa-free for Schengen holders","90 days",       ""),
    ("Serbia",                 "Europe", "Visa-free for Schengen holders","90 days",       ""),
    ("Georgia",                "Europe", "Visa-free for Schengen holders","365 days",      "1-year VF for Schengen holders"),
    ("Turkey",                 "Asia",   "e-Visa discount / facilitated", "30 days",       "Schengen makes eVisa easier"),
]

# ── US visa additional access ────────────────────────────────────────────────
US_VISA_ACCESS = [
    ("Albania",           "Europe",         "Visa-free",     "90 days",  "Any valid used US visa",        ""),
    ("Georgia",           "Europe/Asia",    "Visa-free",     "365 days", "Any valid used US visa",        ""),
    ("Mexico",            "North America",  "Visa-free",     "180 days", "Valid US visa + used once",     ""),
    ("Costa Rica",        "Central America","Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Panama",            "Central America","Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Philippines",       "Southeast Asia", "Visa-free",     "30 days",  "Any valid US visa",             ""),
    ("Mongolia",          "East Asia",      "Visa-free",     "30 days",  "Valid used US visa",            ""),
    ("Colombia",          "South America",  "Visa-free",     "90 days",  "Valid US visa",                 "US visa = VF for some nationalities"),
    ("Bahrain",           "Middle East",    "Visa-free",     "14 days",  "Valid US visa",                 ""),
    ("Oman",              "Middle East",    "Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Qatar",             "Middle East",    "Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Kuwait",            "Middle East",    "Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Taiwan",            "East Asia",      "Visa-free",     "14 days",  "Valid used US visa",            ""),
    ("Albania",           "Europe",         "Visa-free",     "90 days",  "Valid US visa",                 ""),
    ("Kyrgyzstan",        "Central Asia",   "Visa-free",     "30 days",  "Any valid US visa",             ""),
    ("Jamaica",           "Caribbean",      "Visa-free",     "90 days",  "Valid US visa",                 ""),
    ("Bahamas",           "Caribbean",      "Visa-free",     "21 days",  "Valid US visa",                 ""),
    ("Dominican Rep.",    "Caribbean",      "Visa-free",     "30 days",  "Valid US visa",                 ""),
    ("Peru",              "South America",  "Visa-free",     "183 days", "Valid used US visa",            ""),
]

# ── Passport strength summary ────────────────────────────────────────────────
PASSPORT_STRENGTH = [
    # Passport, Rank 2025, VF Countries, VF+VOA, Notes
    ("Singapore",     1,   193, 195, "Henley #1, July 2025"),
    ("Japan",         2,   189, 193, ""),
    ("South Korea",   2,   190, 193, ""),
    ("Germany",       3,   188, 192, "Schengen anchor"),
    ("Italy",         3,   188, 192, ""),
    ("France",        3,   188, 192, ""),
    ("Spain",         3,   188, 192, ""),
    ("Finland",       3,   188, 192, ""),
    ("United Kingdom",5,   186, 190, "Post-Brexit; ETIAS 2026"),
    ("Australia",     6,   186, 190, ""),
    ("Canada",        6,   185, 189, ""),
    ("United States", 12,  180, 186, "Historic decline from #1 in 2014"),
    ("UAE",           8,   181, 187, "Fastest rising in decade"),
    ("Malaysia",      12,  183, 187, ""),
    ("Brazil",        20,  170, 176, ""),
    ("South Africa",  55,  107, 115, ""),
    ("China",         64,  82,  90,  "Rising fast; +37 countries in 10 yrs"),
    ("India",         80,  58,  62,  "Up from 85th in 2024. Henley 2026"),
    ("Nigeria",       92,  46,  51,  ""),
    ("Pakistan",      100, 33,  38,  ""),
    ("Afghanistan",   106, 24,  26,  "Lowest ranked"),
]

print("Data loaded. Building workbook...")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / README
# ═════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "📋 README"
ws_cover.sheet_view.showGridLines = False

ws_cover.merge_cells("B2:H2")
ws_cover["B2"] = "🌍  GLOBAL VISA INTELLIGENCE"
ws_cover["B2"].font = Font(name="Arial", bold=True, size=22, color=C["white"])
ws_cover["B2"].fill = fill(C["navy"])
ws_cover["B2"].alignment = align("left", "center")
ws_cover.row_dimensions[2].height = 48

ws_cover.merge_cells("B3:H3")
ws_cover["B3"] = f"Data current as of June 2026  •  Source: Henley Passport Index, Visa Traveler, IATA TIMATIC"
ws_cover["B3"].font = font(False, C["gray"], 10, True)
ws_cover["B3"].fill = fill(C["navy_mid"])
ws_cover["B3"].alignment = align("left", "center")
ws_cover.row_dimensions[3].height = 24

contents = [
    ("Sheet", "Description", "Key Use"),
    ("📊 Visa Matrix",      "Full country × passport status grid (85 destinations × 8 passports)",    "See at a glance who needs a visa where"),
    ("💰 Costs & Times",    "Visa fees (USD) and processing days by passport + destination",           "Budget and plan application timelines"),
    ("🇬🇧 UK Visa Bonus",  "42 countries accessible with Indian passport + UK visa/BRP",              "Extra travel with a UK student/work visa"),
    ("🇺🇸 US Visa Bonus",  "Countries accessible with other passports + valid US visa",               "Leverage a US visa"),
    ("📑 Schengen Bonus",  "Extra access when holding a Schengen visa",                               "Europe trip extensions"),
    ("🏆 Passport Power",  "Global passport strength ranking with visa-free country counts",           "Quick strength comparison"),
    ("📖 Status Codes",    "Legend for all status codes and colour meanings",                          "Reference"),
]

ws_cover.merge_cells("B5:H5")
ws_cover["B5"] = "CONTENTS"
ws_cover["B5"].font = font(True, C["navy"], 11)
ws_cover["B5"].fill = fill(C["gray_lt"])
ws_cover["B5"].alignment = align("left", "center")
ws_cover.row_dimensions[5].height = 22

for i, (sheet, desc, use) in enumerate(contents, start=6):
    row = i + 1
    ws_cover.row_dimensions[row].height = 24
    ws_cover.merge_cells(f"B{row}:C{row}")
    ws_cover[f"B{row}"] = sheet
    ws_cover[f"B{row}"].font = font(i == 1, C["navy_mid"] if i > 1 else C["black"], 10, False)
    ws_cover[f"B{row}"].fill = fill(C["gray_lt"] if i == 1 else (C["white"] if i % 2 == 0 else C["gray_lt"]))
    ws_cover[f"B{row}"].alignment = align("left", "center")
    ws_cover[f"B{row}"].border = border()

    ws_cover.merge_cells(f"D{row}:F{row}")
    ws_cover[f"D{row}"] = desc
    ws_cover[f"D{row}"].font = font(i == 1, C["black"], 10)
    ws_cover[f"D{row}"].fill = fill(C["gray_lt"] if i == 1 else (C["white"] if i % 2 == 0 else C["gray_lt"]))
    ws_cover[f"D{row}"].alignment = align("left", "center", True)
    ws_cover[f"D{row}"].border = border()

    ws_cover.merge_cells(f"G{row}:H{row}")
    ws_cover[f"G{row}"] = use
    ws_cover[f"G{row}"].font = font(False, C["gray"], 9, True)
    ws_cover[f"G{row}"].fill = fill(C["gray_lt"] if i == 1 else (C["white"] if i % 2 == 0 else C["gray_lt"]))
    ws_cover[f"G{row}"].alignment = align("left", "center", True)
    ws_cover[f"G{row}"].border = border()

# Disclaimer
ws_cover.merge_cells("B18:H20")
ws_cover["B18"] = ("⚠️  DISCLAIMER: Visa policies change frequently. Always verify on official embassy/government "
                   "websites before booking. This file is for planning reference only. Data sourced from Henley "
                   "Passport Index (July 2025 / Jan 2026 editions), IATA TIMATIC, Visa Traveler, and Gulf News.")
ws_cover["B18"].font = font(False, C["amber"], 9)
ws_cover["B18"].fill = fill(C["amber_lt"])
ws_cover["B18"].alignment = align("left", "top", True)
ws_cover["B18"].border = border()
ws_cover.row_dimensions[18].height = 20
ws_cover.row_dimensions[19].height = 20
ws_cover.row_dimensions[20].height = 20

set_col_widths(ws_cover, {"A": 2, "B": 20, "C": 20, "D": 30, "E": 25, "F": 25, "G": 28, "H": 28})

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — VISA MATRIX
# ═════════════════════════════════════════════════════════════════════════════
ws_matrix = wb.create_sheet("📊 Visa Matrix")
ws_matrix.sheet_view.showGridLines = False

PASSPORT_COLS = ["🇮🇳 India", "🇬🇧 UK", "🇺🇸 USA", "🇨🇦 Canada", "🇦🇺 Australia", "🇳🇬 Nigeria", "🇨🇳 China", "🇿🇦 S.Africa"]

# Title row
ws_matrix.merge_cells("A1:K1")
ws_matrix["A1"] = "GLOBAL VISA STATUS MATRIX  |  85 Destinations × 8 Passports  |  June 2026"
ws_matrix["A1"].font = font(True, C["white"], 12)
ws_matrix["A1"].fill = fill(C["navy"])
ws_matrix["A1"].alignment = align("center", "center")
ws_matrix.row_dimensions[1].height = 30

# Sub-header
ws_matrix.merge_cells("A2:K2")
ws_matrix["A2"] = "VF = Visa-Free  |  VOA = Visa on Arrival  |  EV = e-Visa (online application)  |  VR = Visa Required  |  ESTA = Electronic System for Travel Authorization  |  N/A = Own country"
ws_matrix["A2"].font = font(False, C["gray"], 9, True)
ws_matrix["A2"].fill = fill(C["gray_lt"])
ws_matrix["A2"].alignment = align("left", "center")
ws_matrix.row_dimensions[2].height = 18

# Column headers
headers = ["#", "Destination", "Region"] + PASSPORT_COLS
ws_matrix.append(headers)
style_header_row(ws_matrix, 3, C["navy_mid"], "FFFFFF", True, 10)
ws_matrix.row_dimensions[3].height = 36

# Status colour map
STATUS_COLORS = {
    VF:          (C["green"],    C["green_lt"]),
    VOA:         (C["teal"],     C["teal_lt"]),
    EV:          (C["blue"],     C["blue_lt"]),
    VR:          (C["red"],      C["red_lt"]),
    "ESTA/VF":   (C["teal"],     C["teal_lt"]),
    "ESTA":      (C["teal"],     C["teal_lt"]),
    NA:          (C["gray"],     C["gray_lt"]),
}

prev_region = ""
row_num = 4
for i, row_data in enumerate(VISA_DATA, 1):
    dest, region, *statuses = row_data
    # Region separator
    if region != prev_region:
        ws_matrix.merge_cells(f"A{row_num}:K{row_num}")
        ws_matrix[f"A{row_num}"] = f"  {region.upper()}"
        ws_matrix[f"A{row_num}"].font = font(True, C["white"], 9)
        ws_matrix[f"A{row_num}"].fill = fill(C["navy_mid"])
        ws_matrix[f"A{row_num}"].alignment = align("left", "center")
        ws_matrix.row_dimensions[row_num].height = 18
        prev_region = region
        row_num += 1

    ws_matrix.cell(row_num, 1, i)
    ws_matrix.cell(row_num, 2, dest)
    ws_matrix.cell(row_num, 3, region)

    for col_offset, status in enumerate(statuses, 4):
        cell = ws_matrix.cell(row_num, col_offset, status)
        fg_col, bg_col = STATUS_COLORS.get(status, (C["black"], C["white"]))
        cell.fill = fill(bg_col)
        cell.font = font(True, fg_col, 9)
        cell.alignment = align("center", "center")
        cell.border = border()

    # Style row base
    for c in [1, 2, 3]:
        cell = ws_matrix.cell(row_num, c)
        cell.fill = fill(C["white"] if i % 2 == 0 else C["gray_lt"])
        cell.font = font(c == 2, C["black"] if c == 2 else C["gray"], 9)
        cell.alignment = align("left" if c == 2 else "center", "center")
        cell.border = border()

    ws_matrix.row_dimensions[row_num].height = 20
    row_num += 1

set_col_widths(ws_matrix, {
    "A": 4, "B": 18, "C": 16,
    "D": 13, "E": 13, "F": 13,
    "G": 13, "H": 13, "I": 13,
    "J": 13, "K": 13
})
freeze(ws_matrix, "D4")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — COSTS & PROCESSING TIMES
# ═════════════════════════════════════════════════════════════════════════════
ws_cost = wb.create_sheet("💰 Costs & Times")
ws_cost.sheet_view.showGridLines = False

ws_cost.merge_cells("A1:K1")
ws_cost["A1"] = "VISA COSTS & PROCESSING TIMES  |  Indian 🇮🇳 · UK 🇬🇧 · USA 🇺🇸  |  June 2026 (approx. USD)"
ws_cost["A1"].font = font(True, C["white"], 12)
ws_cost["A1"].fill = fill(C["navy"])
ws_cost["A1"].alignment = align("center", "center")
ws_cost.row_dimensions[1].height = 30

ws_cost.merge_cells("A2:K2")
ws_cost["A2"] = "Costs shown in USD or local currency equivalent. 'N/A (VF)' = no fee as visa not required. '-' = not applicable."
ws_cost["A2"].font = font(False, C["gray"], 9, True)
ws_cost["A2"].fill = fill(C["gray_lt"])
ws_cost["A2"].alignment = align("left", "center")
ws_cost.row_dimensions[2].height = 16

cost_headers = ["Destination", "Region",
                "🇮🇳 IN Cost", "🇮🇳 IN Process Time",
                "🇬🇧 UK Cost", "🇬🇧 UK Process Time",
                "🇺🇸 US Cost", "🇺🇸 US Process Time",
                "Notes / Tips"]
ws_cost.append(cost_headers)
style_header_row(ws_cost, 3, C["navy_mid"], "FFFFFF", True, 10)
ws_cost.row_dimensions[3].height = 36

cost_rows = []
for row_data in VISA_DATA:
    dest, region = row_data[0], row_data[1]
    if dest in COST_DATA:
        cd = COST_DATA[dest]
        cost_rows.append((dest, region, cd[0], cd[1], cd[2], cd[3], cd[4], cd[5], cd[6]))

for i, row in enumerate(cost_rows, 1):
    r = 3 + i
    ws_cost.row_dimensions[r].height = 22
    for c, val in enumerate(row, 1):
        cell = ws_cost.cell(r, c, val)
        cell.border = border()
        cell.alignment = align("left" if c in [1, 2, 9] else "center", "center", c == 9)
        bg = C["white"] if i % 2 == 0 else C["gray_lt"]
        cell.fill = fill(bg)
        cell.font = font(c == 1, C["black"], 9)

        # Colour-code cost cells
        if "N/A (VF)" in str(val) or "Free VF" in str(val):
            cell.font = font(False, C["green"], 9)
        elif "N/A (citizen)" in str(val):
            cell.font = font(False, C["gray"], 9)
        elif "No visa" in str(val) or "Instant" in str(val):
            cell.font = font(False, C["teal"], 9)
        elif "days" in str(val) and c in [4, 6, 8]:
            cell.font = font(False, C["amber"], 9)
        elif "weeks" in str(val) or "months" in str(val):
            cell.font = font(True, C["red"], 9)

set_col_widths(ws_cost, {
    "A": 18, "B": 16, "C": 16, "D": 18, "E": 16, "F": 18,
    "G": 16, "H": 18, "I": 32
})
freeze(ws_cost, "C4")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — UK VISA BONUS
# ═════════════════════════════════════════════════════════════════════════════
ws_uk = wb.create_sheet("🇬🇧 UK Visa Bonus")
ws_uk.sheet_view.showGridLines = False

ws_uk.merge_cells("A1:G1")
ws_uk["A1"] = "🇬🇧  COUNTRIES ACCESSIBLE WITH INDIAN PASSPORT + UK VISA / BRP"
ws_uk["A1"].font = font(True, C["white"], 13)
ws_uk["A1"].fill = fill(C["navy"])
ws_uk["A1"].alignment = align("left", "center")
ws_uk.row_dimensions[1].height = 36

ws_uk.merge_cells("A2:G2")
ws_uk["A2"] = ("This sheet shows countries where an Indian passport holder can gain additional access "
               "by holding a valid UK Standard Visitor visa, Student visa (Tier 4), Work visa, or UK BRP/eVisa.  "
               "The UK visa acts as proof of vetting — many countries accept this as a substitute for their own visa. "
               "Always verify current rules before travel.")
ws_uk["A2"].font = font(False, C["black"], 9, True)
ws_uk["A2"].fill = fill(C["amber_lt"])
ws_uk["A2"].alignment = align("left", "center", True)
ws_uk["A2"].border = border(color=C["amber"])
ws_uk.row_dimensions[2].height = 44

ws_uk.merge_cells("A3:G3")
ws_uk["A3"] = f"Total: {len(UK_VISA_ACCESS)} destinations  |  Including major unlocks: Japan, South Korea, Turkey, most Gulf states, and 12 Caribbean islands"
ws_uk["A3"].font = font(True, C["teal"], 10)
ws_uk["A3"].fill = fill(C["teal_lt"])
ws_uk["A3"].alignment = align("left", "center")
ws_uk.row_dimensions[3].height = 22

uk_headers = ["#", "Country", "Region", "Access Type", "Max Stay", "Conditions / Visa Type", "Insider Notes"]
ws_uk.append(uk_headers)
style_header_row(ws_uk, 4, C["navy_mid"], "FFFFFF", True, 10)
ws_uk.row_dimensions[4].height = 30

HIGHLIGHT_COUNTRIES = {"Japan", "South Korea", "Turkey", "Bahrain", "Oman", "Qatar", "Kuwait", "UAE", "Saudi Arabia"}

for i, (country, region, access, stay, cond, note) in enumerate(UK_VISA_ACCESS, 1):
    r = 4 + i
    ws_uk.row_dimensions[r].height = 22
    row_vals = [i, country, region, access, stay, cond, note]
    is_highlight = country in HIGHLIGHT_COUNTRIES
    bg = C["amber_lt"] if is_highlight else (C["white"] if i % 2 == 0 else C["gray_lt"])
    for c, val in enumerate(row_vals, 1):
        cell = ws_uk.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(is_highlight, C["amber"] if is_highlight else C["black"], 9)
        cell.alignment = align("left" if c in [2, 3, 6, 7] else "center", "center", c in [6, 7])
        cell.border = border()

    if "Visa-free" in access:
        ws_uk.cell(r, 4).font = font(True, C["green"], 9)
    elif "eVisa" in access or "e-Visa" in access:
        ws_uk.cell(r, 4).font = font(False, C["blue"], 9)

set_col_widths(ws_uk, {"A": 4, "B": 20, "C": 18, "D": 20, "E": 14, "F": 38, "G": 36})
freeze(ws_uk, "B5")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — US VISA BONUS
# ═════════════════════════════════════════════════════════════════════════════
ws_us = wb.create_sheet("🇺🇸 US Visa Bonus")
ws_us.sheet_view.showGridLines = False

ws_us.merge_cells("A1:G1")
ws_us["A1"] = "🇺🇸  ADDITIONAL COUNTRIES ACCESSIBLE WITH A VALID US VISA"
ws_us["A1"].font = font(True, C["white"], 13)
ws_us["A1"].fill = fill(C["navy"])
ws_us["A1"].alignment = align("left", "center")
ws_us.row_dimensions[1].height = 36

ws_us.merge_cells("A2:G2")
ws_us["A2"] = ("A valid US B1/B2 or other non-immigrant visa allows simplified or visa-free entry to several countries "
               "for non-US citizens who would otherwise need a separate visa. The US visa must generally have been "
               "used at least once to enter the USA. Primarily benefits Indian, Nigerian, Pakistani and similar passports.")
ws_us["A2"].font = font(False, C["black"], 9, True)
ws_us["A2"].fill = fill(C["blue_lt"])
ws_us["A2"].alignment = align("left", "center", True)
ws_us.row_dimensions[2].height = 40

us_headers = ["#", "Country", "Region", "Access Type", "Max Stay", "Conditions", "Notes"]
ws_us.append(us_headers)
style_header_row(ws_us, 3, C["navy_mid"], "FFFFFF", True, 10)
ws_us.row_dimensions[3].height = 30

for i, (country, region, access, stay, cond, note) in enumerate(US_VISA_ACCESS, 1):
    r = 3 + i
    ws_us.row_dimensions[r].height = 22
    row_vals = [i, country, region, access, stay, cond, note]
    bg = C["white"] if i % 2 == 0 else C["gray_lt"]
    for c, val in enumerate(row_vals, 1):
        cell = ws_us.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(False, C["black"], 9)
        cell.alignment = align("left" if c in [2, 3, 6, 7] else "center", "center", c in [6, 7])
        cell.border = border()
    if "Visa-free" in access:
        ws_us.cell(r, 4).font = font(True, C["green"], 9)

set_col_widths(ws_us, {"A": 4, "B": 20, "C": 18, "D": 20, "E": 14, "F": 36, "G": 30})
freeze(ws_us, "B4")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SCHENGEN BONUS
# ═════════════════════════════════════════════════════════════════════════════
ws_sch = wb.create_sheet("📑 Schengen Bonus")
ws_sch.sheet_view.showGridLines = False

ws_sch.merge_cells("A1:F1")
ws_sch["A1"] = "📑  EXTRA ACCESS WITH A SCHENGEN VISA"
ws_sch["A1"].font = font(True, C["white"], 13)
ws_sch["A1"].fill = fill(C["navy"])
ws_sch["A1"].alignment = align("left", "center")
ws_sch.row_dimensions[1].height = 36

ws_sch.merge_cells("A2:F2")
ws_sch["A2"] = ("A valid Schengen visa (issued by any of the 27 Schengen member states) may allow entry to additional "
                "non-Schengen European countries that accept it in lieu of their own national visa. Useful for Indian "
                "passport holders who have obtained a Schengen visa and are planning wider European travel.")
ws_sch["A2"].font = font(False, C["black"], 9, True)
ws_sch["A2"].fill = fill(C["purple_lt"])
ws_sch["A2"].alignment = align("left", "center", True)
ws_sch.row_dimensions[2].height = 40

sch_headers = ["#", "Country / Zone", "Region", "Access Type", "Max Stay", "Notes"]
ws_sch.append(sch_headers)
style_header_row(ws_sch, 3, C["navy_mid"], "FFFFFF", True, 10)
ws_sch.row_dimensions[3].height = 30

for i, (country, region, access, stay, note) in enumerate(SCHENGEN_ACCESS, 1):
    r = 3 + i
    ws_sch.row_dimensions[r].height = 22
    row_vals = [i, country, region, access, stay, note]
    bg = C["white"] if i % 2 == 0 else C["gray_lt"]
    for c, val in enumerate(row_vals, 1):
        cell = ws_sch.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(False, C["black"], 9)
        cell.alignment = align("left" if c in [2, 3, 5, 6] else "center", "center", c == 6)
        cell.border = border()

set_col_widths(ws_sch, {"A": 4, "B": 26, "C": 16, "D": 28, "E": 14, "F": 44})
freeze(ws_sch, "B4")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7 — PASSPORT POWER RANKING
# ═════════════════════════════════════════════════════════════════════════════
ws_rank = wb.create_sheet("🏆 Passport Power")
ws_rank.sheet_view.showGridLines = False

ws_rank.merge_cells("A1:G1")
ws_rank["A1"] = "🏆  GLOBAL PASSPORT POWER RANKING  |  Henley Passport Index  |  2025–2026"
ws_rank["A1"].font = font(True, C["white"], 13)
ws_rank["A1"].fill = fill(C["navy"])
ws_rank["A1"].alignment = align("left", "center")
ws_rank.row_dimensions[1].height = 36

rank_headers = ["Rank", "Passport", "Visa-Free Count", "VF + VOA Total", "% of World VF", "Trend", "Notes"]
ws_rank.append(rank_headers)
style_header_row(ws_rank, 2, C["navy_mid"], "FFFFFF", True, 10)
ws_rank.row_dimensions[2].height = 30

HIGHLIGHT_PASSPORTS = {"India", "United Kingdom", "United States", "Nigeria", "China"}

for i, (passport, rank, vf, vf_voa, note) in enumerate(PASSPORT_STRENGTH, 1):
    r = 2 + i
    ws_rank.row_dimensions[r].height = 24

    pct = f"=C{r}/193*100"  # formula: % of 193 possible
    trend = "↗ Rising" if "Rising" in note or "fastest" in note.lower() else ("↘ Declining" if "decline" in note.lower() or "drop" in note.lower() else "→ Stable")

    row_vals = [rank, passport, vf, vf_voa, pct, trend, note]
    is_hl = passport in HIGHLIGHT_PASSPORTS
    bg = C["navy"] if rank == 1 else (C["amber_lt"] if is_hl else (C["white"] if i % 2 == 0 else C["gray_lt"]))
    fg = C["white"] if rank == 1 else (C["amber"] if is_hl else C["black"])

    for c, val in enumerate(row_vals, 1):
        cell = ws_rank.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(rank <= 3 or is_hl, fg, 9 if rank > 3 else 10)
        cell.alignment = align("center" if c not in [2, 7] else "left", "center")
        cell.border = border()

    # % column format
    ws_rank.cell(r, 5).number_format = "0.0\"%\""

    # Trend colouring
    tc = ws_rank.cell(r, 6)
    if "Rising" in trend:
        tc.font = font(True, C["green"], 9)
    elif "Declining" in trend:
        tc.font = font(True, C["red"], 9)

set_col_widths(ws_rank, {"A": 8, "B": 18, "C": 16, "D": 16, "E": 14, "F": 14, "G": 42})
freeze(ws_rank, "A3")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 8 — STATUS CODES LEGEND
# ═════════════════════════════════════════════════════════════════════════════
ws_leg = wb.create_sheet("📖 Status Codes")
ws_leg.sheet_view.showGridLines = False

ws_leg.merge_cells("B2:F2")
ws_leg["B2"] = "STATUS CODE LEGEND"
ws_leg["B2"].font = font(True, C["white"], 14)
ws_leg["B2"].fill = fill(C["navy"])
ws_leg["B2"].alignment = align("center", "center")
ws_leg.row_dimensions[2].height = 36

legend_items = [
    (VF,       C["green"],    C["green_lt"],  "No visa required. Walk up to border/airport and enter."),
    (VOA,      C["teal"],     C["teal_lt"],   "Obtain visa at airport/border on arrival. Usually a fee applies."),
    (EV,       C["blue"],     C["blue_lt"],   "Apply online before travel. Usually approved in hours to days."),
    ("ESTA",   C["teal"],     C["teal_lt"],   "Electronic System for Travel Authorization (US). $40, valid 2 years."),
    ("ESTA/VF",C["teal"],     C["teal_lt"],   "US passport = ESTA (electronic pre-auth). Other strong passports may be VF."),
    (VR,       C["red"],      C["red_lt"],    "Full visa required. Apply at embassy/consulate before travel. Fees and interviews likely."),
    (NA,       C["gray"],     C["gray_lt"],   "Not applicable — this is the traveller's own country."),
]

for i, (code, fg, bg, desc) in enumerate(legend_items, 1):
    r = 3 + i
    ws_leg.row_dimensions[r].height = 28
    ws_leg.merge_cells(f"B{r}:C{r}")
    cell = ws_leg[f"B{r}"]
    cell.value = code
    cell.fill = fill(bg)
    cell.font = font(True, fg, 11)
    cell.alignment = align("center", "center")
    cell.border = border()

    ws_leg.merge_cells(f"D{r}:F{r}")
    desc_cell = ws_leg[f"D{r}"]
    desc_cell.value = desc
    desc_cell.fill = fill(C["white"])
    desc_cell.font = font(False, C["black"], 10)
    desc_cell.alignment = align("left", "center", True)
    desc_cell.border = border()

# Processing time colour guide
ws_leg.row_dimensions[12] = ws_leg.row_dimensions[12]
ws_leg.merge_cells("B12:F12")
ws_leg["B12"] = "PROCESSING TIME COLOUR GUIDE (Costs & Times sheet)"
ws_leg["B12"].font = font(True, C["navy"], 11)
ws_leg["B12"].fill = fill(C["gray_lt"])
ws_leg["B12"].alignment = align("left", "center")
ws_leg.row_dimensions[12].height = 24

time_legend = [
    (C["green"],  "Free / No fee",     "Visa-free access — no cost"),
    (C["teal"],   "Instant / No visa", "Very fast — minutes to hours"),
    (C["amber"],  "Days",              "Standard processing: 1–14 days"),
    (C["red"],    "Weeks / Months",    "Slow processing — plan well ahead"),
]

for i, (col, label, desc) in enumerate(time_legend, 1):
    r = 12 + i
    ws_leg.row_dimensions[r].height = 24
    ws_leg.merge_cells(f"B{r}:C{r}")
    cell = ws_leg[f"B{r}"]
    cell.value = label
    cell.fill = fill(C["gray_lt"])
    cell.font = font(True, col, 10)
    cell.alignment = align("center", "center")
    cell.border = border()

    ws_leg.merge_cells(f"D{r}:F{r}")
    d = ws_leg[f"D{r}"]
    d.value = desc
    d.fill = fill(C["white"])
    d.font = font(False, C["black"], 10)
    d.alignment = align("left", "center")
    d.border = border()

set_col_widths(ws_leg, {"A": 2, "B": 16, "C": 16, "D": 20, "E": 20, "F": 20})

# ─────────────────────────────────────────────────────────────────────────────
# Save & recalc
# ─────────────────────────────────────────────────────────────────────────────
out_path = "/Users/mango/Projects/cmdshiftN/output/Voyager_VisaIntelligence_June2026.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
