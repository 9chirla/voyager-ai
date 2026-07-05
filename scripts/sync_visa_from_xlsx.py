#!/usr/bin/env python3
"""Sync visa data from Voyager_VisaIntelligence_June2026.xlsx into src/data/*.js"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

from visa_reference_data import (
    IN_COST_OVERRIDES,
    IN_NOTES_OVERRIDES,
    OTHER_PASSPORT_CORRECTIONS,
    PASSPORT_REFERENCES,
)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Voyager_VisaIntelligence_June2026.xlsx"
OUT_MATRIX = ROOT / "src/data/visaMatrix.js"
OUT_COSTS = ROOT / "src/data/visaCosts.js"
OUT_HELD = ROOT / "src/data/heldVisaData.js"

NAME_TO_ISO_PATH = ROOT / "scripts/country_name_to_iso.json"


def load_name_to_iso() -> dict:
    with open(NAME_TO_ISO_PATH, encoding="utf-8") as f:
        return json.load(f)

PASSPORT_COLS = {
    "India": "IN",
    "UK": "GB",
    "USA": "US",
    "Canada": "CA",
    "Australia": "AU",
    "Nigeria": "NG",
    "China": "CN",
    "S.Africa": "ZA",
}

COST_PASSPORT_COLS = {"IN": (2, 3), "GB": (4, 5), "US": (6, 7)}

# Passports in the selector but absent from the xlsx matrix columns
SUPPLEMENTAL_VISA_FREE = {
    "JP": ["FR", "ES", "IT", "PT", "GR", "HR", "PL", "TH", "MY", "SG", "KR", "JO", "MA", "MX", "CR", "CO", "AE", "VN", "ID", "PH", "LK"],
    "BR": ["FR", "ES", "IT", "PT", "GR", "HR", "PL", "TH", "MY", "SG", "JO", "MA", "KE", "ZA", "MX", "CR", "CO", "AE", "VN", "ID", "PH", "LK"],
    "MY": ["FR", "ES", "IT", "PT", "GR", "HR", "PL", "TH", "SG", "JP", "KR", "JO", "MA", "KE", "ZA", "MX", "CR", "CO", "AE", "VN", "ID", "PH", "LK"],
}
SUPPLEMENTAL_VOA = {
    "JP": ["KE", "ZA", "AU", "NZ"],
    "BR": ["KE", "TH", "VN"],
    "MY": ["KE"],
}
SUPPLEMENTAL_E_VISA = {
    "JP": ["VN", "AU", "NZ"],
    "BR": ["AU", "NZ", "JP", "KR"],
    "MY": ["AU", "NZ"],
}

def filter_dest_lists(data: dict) -> dict:
    """Drop self-references (passport country = destination)."""
    out: dict[str, list[str]] = {}
    for passport, dests in data.items():
        filtered = sorted({d for d in dests if d != passport})
        if filtered:
            out[passport] = filtered
    return out


def merge_supplemental(visa_free: dict, voa: dict, evisa: dict) -> None:
    for passport, dests in SUPPLEMENTAL_VISA_FREE.items():
        merged = sorted(set(visa_free.get(passport, []) + dests))
        visa_free[passport] = merged
    for passport, dests in SUPPLEMENTAL_VOA.items():
        merged = sorted(set(voa.get(passport, []) + dests))
        if merged:
            voa[passport] = merged
    for passport, dests in SUPPLEMENTAL_E_VISA.items():
        merged = sorted(set(evisa.get(passport, []) + dests))
        if merged:
            evisa[passport] = merged


def normalize_status(raw: str | None, passport_iso: str, dest_iso: str) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.startswith("  "):  # region header rows
        return None
    if passport_iso == dest_iso or "N/A" in s and "Own" in s:
        return "visa-free"
    if "Visa-Free" in s or s == "ESTA/VF":
        return "visa-free"
    if "Visa on Arrival" in s or s == "VOA":
        return "visa-on-arrival"
    if "e-Visa" in s or s == "EV" or s == "ESTA":
        return "e-visa"
    if "Visa Required" in s or s == "VR":
        return "visa-required"
    return None


def iso_for_name(name: str, name_to_iso: dict) -> str | None:
    if not name:
        return None
    name = str(name).strip()
    if name in name_to_iso:
        return name_to_iso[name]
    if "Schengen" in name:
        return None
    return None


def parse_matrix_sheet(wb, name_to_iso: dict) -> tuple[dict, dict, dict]:
    ws = wb["📊 Visa Matrix"]
    rows = list(ws.iter_rows(values_only=True))
    header = None
    passport_isos: list[str] = []

    visa_free: dict[str, list[str]] = {}
    voa: dict[str, list[str]] = {}
    evisa: dict[str, list[str]] = {}

    for row in rows:
        if row[0] == "#":
            header = row
            for col in row[3:]:
                if not col:
                    continue
                label = re.sub(r"^[\U0001F1E6-\U0001F1FF]{2}\s*", "", str(col)).strip()
                for key, iso in PASSPORT_COLS.items():
                    if key in label:
                        passport_isos.append(iso)
                        break
            continue
        if not header or not row[1] or not isinstance(row[0], (int, float)):
            continue
        dest_name = str(row[1]).strip()
        dest_iso = iso_for_name(dest_name, name_to_iso)
        if not dest_iso:
            continue

        for i, passport_iso in enumerate(passport_isos):
            status = normalize_status(row[3 + i], passport_iso, dest_iso)
            if not status or status == "visa-required":
                continue
            bucket = {"visa-free": visa_free, "visa-on-arrival": voa, "e-visa": evisa}[status]
            bucket.setdefault(passport_iso, [])
            if dest_iso not in bucket[passport_iso]:
                bucket[passport_iso].append(dest_iso)

    # EU selector mirrors Schengen member access (same as GB for our EU destinations)
    if "GB" in visa_free:
        visa_free["EU"] = sorted(set(visa_free["GB"]))
    if "GB" in voa:
        voa["EU"] = sorted(set(voa["GB"]))
    if "GB" in evisa:
        evisa["EU"] = sorted(set(evisa["GB"]))

    visa_free = filter_dest_lists(visa_free)
    voa = filter_dest_lists(voa)
    evisa = filter_dest_lists(evisa)
    merge_supplemental(visa_free, voa, evisa)
    apply_passport_references(visa_free, voa, evisa)
    apply_matrix_corrections(visa_free, voa, evisa, OTHER_PASSPORT_CORRECTIONS)

    return visa_free, voa, evisa


def apply_passport_references(
    visa_free: dict[str, list[str]],
    voa: dict[str, list[str]],
    evisa: dict[str, list[str]],
) -> None:
    """Rebuild passport buckets from verified reference data (overrides xlsx)."""
    buckets = {
        "visa-free": visa_free,
        "visa-on-arrival": voa,
        "e-visa": evisa,
    }

    for passport_iso, reference in PASSPORT_REFERENCES.items():
        for bucket in buckets.values():
            bucket.pop(passport_iso, None)

        for dest_iso, status in reference.items():
            if status == "visa-required":
                continue
            bucket = buckets[status]
            dests = bucket.setdefault(passport_iso, [])
            if dest_iso not in dests:
                dests.append(dest_iso)

        for bucket in buckets.values():
            if passport_iso in bucket:
                bucket[passport_iso] = sorted(set(bucket[passport_iso]))


def apply_matrix_corrections(
    visa_free: dict[str, list[str]],
    voa: dict[str, list[str]],
    evisa: dict[str, list[str]],
    corrections: dict[tuple[str, str], str],
) -> None:
    buckets = {
        "visa-free": visa_free,
        "visa-on-arrival": voa,
        "e-visa": evisa,
    }

    for (passport, dest), status in corrections.items():
        if passport in PASSPORT_REFERENCES:
            continue
        for bucket in buckets.values():
            if passport in bucket:
                bucket[passport] = [d for d in bucket[passport] if d != dest]
                if not bucket[passport]:
                    del bucket[passport]

        if status in buckets:
            dests = buckets[status].setdefault(passport, [])
            if dest not in dests:
                dests.append(dest)
                dests.sort()

    for bucket in buckets.values():
        for passport in list(bucket.keys()):
            bucket[passport] = sorted(set(bucket[passport]))


def reconcile_in_costs(
    visa_free: dict[str, list[str]],
    voa: dict[str, list[str]],
    evisa: dict[str, list[str]],
    costs: dict[str, dict],
) -> None:
    """Align Indian cost rows with verified passport-alone statuses."""

    def status_for(dest_iso: str) -> str:
        if dest_iso in visa_free.get("IN", []):
            return "visa-free"
        if dest_iso in voa.get("IN", []):
            return "visa-on-arrival"
        if dest_iso in evisa.get("IN", []):
            return "e-visa"
        return "visa-required"

    for dest_iso, entry in costs.items():
        if dest_iso in IN_NOTES_OVERRIDES:
            entry["notes"] = IN_NOTES_OVERRIDES[dest_iso]

        st = status_for(dest_iso)
        if dest_iso in IN_COST_OVERRIDES:
            entry["IN"] = IN_COST_OVERRIDES[dest_iso]
            continue

        in_entry = entry.get("IN")
        if not in_entry:
            continue
        if st != "visa-free" and in_entry.get("cost") == "Free VF":
            if st == "e-visa":
                entry["IN"] = {"cost": "e-Visa required", "processing": "Apply online"}
            elif st == "visa-on-arrival":
                entry["IN"] = {"cost": "VOA fee applies", "processing": "On arrival"}
            else:
                entry["IN"] = {"cost": "Embassy visa", "processing": "Varies"}


def parse_costs_sheet(wb, name_to_iso: dict) -> dict:
    ws = wb["💰 Costs & Times"]
    costs: dict[str, dict] = {}

    for row in ws.iter_rows(values_only=True):
        if not row[0] or row[0] == "Destination":
            continue
        dest_name = str(row[0]).strip()
        if dest_name.startswith("VISA") or dest_name.startswith("Costs"):
            continue
        dest_iso = iso_for_name(dest_name, name_to_iso)
        if not dest_iso:
            continue

        entry = {"notes": (row[8] or "").strip() if len(row) > 8 else ""}
        for passport_iso, (cost_idx, time_idx) in COST_PASSPORT_COLS.items():
            cost = row[cost_idx] if len(row) > cost_idx else None
            proc = row[time_idx] if len(row) > time_idx else None
            if cost or proc:
                entry[passport_iso] = {
                    "cost": str(cost).strip() if cost else None,
                    "processing": str(proc).strip() if proc else None,
                }
        costs[dest_iso] = entry

    return costs


def grant_type(access: str) -> str | None:
    a = access.lower()
    if "visa-free" in a:
        return "visaFree"
    if "visa on arrival" in a or "voa" in a:
        return "visaOnArrival"
    if "e-visa" in a or "evisa" in a:
        return "eVisa"
    return None


def parse_bonus_sheet(wb, sheet_name: str, grants: dict, name_to_iso: dict) -> None:
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if not row[1] or row[1] == "Country" or row[1] == "Country / Zone":
            continue
        if isinstance(row[0], str) and not str(row[0]).isdigit():
            continue
        country = str(row[1]).strip()
        dest_iso = iso_for_name(country, name_to_iso)
        if not dest_iso:
            continue
        access = str(row[3] or "")
        gtype = grant_type(access)
        if not gtype:
            continue
        grants.setdefault(gtype, [])
        if dest_iso not in grants[gtype]:
            grants[gtype].append(dest_iso)


def js_string(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


def write_matrix(visa_free, voa, evisa) -> None:
    lines = [
        "// AUTO-GENERATED from Voyager_VisaIntelligence_June2026.xlsx — do not edit by hand",
        "// Run: scripts/.venv/bin/python scripts/sync_visa_from_xlsx.py",
        "",
        "/** @type {Record<string, string[]>} Passport ISO → visa-free destination ISOs */",
        f"export const VISA_FREE = {json.dumps(visa_free, indent=2)};",
        "",
        "/** @type {Record<string, string[]>} */",
        f"export const VISA_ON_ARRIVAL = {json.dumps(voa, indent=2)};",
        "",
        "/** @type {Record<string, string[]>} */",
        f"export const E_VISA = {json.dumps(evisa, indent=2)};",
        "",
        """/**
 * @param {string} passportIso
 * @param {string} destinationIso
 * @returns {'visa-free'|'visa-on-arrival'|'e-visa'|'visa-required'|'unknown'}
 */
export function lookupVisaStatus(passportIso, destinationIso) {
  if (passportIso === destinationIso) return 'visa-free';
  if (VISA_FREE[passportIso]?.includes(destinationIso)) return 'visa-free';
  if (VISA_ON_ARRIVAL[passportIso]?.includes(destinationIso)) return 'visa-on-arrival';
  if (E_VISA[passportIso]?.includes(destinationIso)) return 'e-visa';
  if (VISA_FREE[passportIso] || VISA_ON_ARRIVAL[passportIso] || E_VISA[passportIso]) {
    return 'visa-required';
  }
  return 'unknown';
}
""",
    ]
    OUT_MATRIX.write_text("\n".join(lines), encoding="utf-8")


def write_costs(costs: dict) -> None:
    lines = [
        "// AUTO-GENERATED from Voyager_VisaIntelligence_June2026.xlsx",
        "",
        "/** @typedef {{ cost: string|null, processing: string|null }} VisaCostEntry */",
        "/** @typedef {{ notes?: string, IN?: VisaCostEntry, GB?: VisaCostEntry, US?: VisaCostEntry }} DestinationVisaCosts */",
        "",
        "/** @type {Record<string, DestinationVisaCosts>} */",
        f"export const VISA_COSTS = {json.dumps(costs, indent=2, ensure_ascii=False)};",
        "",
        """/**
 * @param {string} destinationIso
 * @param {string} passportIso
 * @returns {VisaCostEntry|null}
 */
export function lookupVisaCost(destinationIso, passportIso) {
  return VISA_COSTS[destinationIso]?.[passportIso] ?? null;
}

/**
 * @param {string} destinationIso
 * @returns {string}
 */
export function lookupVisaNotes(destinationIso) {
  return VISA_COSTS[destinationIso]?.notes ?? '';
}
""",
    ]
    OUT_COSTS.write_text("\n".join(lines), encoding="utf-8")


def write_held(schengen, uk, us, canada, uae, australia) -> None:
    options = [
        {
            "id": "schengen-visa",
            "label": "Schengen visa",
            "flagIso": "EU",
            "hint": "Valid short-stay Schengen visa — covers all 27 Schengen states in our EU destination list.",
            "grants": schengen,
        },
        {
            "id": "uk-visa",
            "label": "UK visa / BRP",
            "flagIso": "GB",
            "hint": "Valid UK visitor, student, or work visa — unlocks Japan, South Korea, Mexico and more for Indian passport holders.",
            "grants": uk,
        },
        {
            "id": "us-visa",
            "label": "US visa",
            "flagIso": "US",
            "hint": "Valid US B1/B2 (used at least once) — visa-free entry to Mexico, Colombia, Costa Rica and more.",
            "grants": us,
        },
        {
            "id": "canada-visa",
            "label": "Canada visa",
            "flagIso": "CA",
            "hint": "Valid Canadian visitor or study permit.",
            "grants": canada,
        },
        {
            "id": "uae-residence",
            "label": "UAE residence",
            "flagIso": "AE",
            "hint": "Valid UAE residence visa or Emirates ID.",
            "grants": uae,
        },
        {
            "id": "australia-visa",
            "label": "Australia visa",
            "flagIso": "AU",
            "hint": "Valid Australian visitor or student visa.",
            "grants": australia,
        },
    ]

    lines = [
        "// AUTO-GENERATED from Voyager_VisaIntelligence_June2026.xlsx",
        "",
        "/**",
        " * Optional held visas / residence permits that change destination access",
        " * beyond what a passport alone grants.",
        " *",
        " * @typedef {object} HeldVisaOption",
        " * @property {string} id",
        " * @property {string} label",
        " * @property {string} flagIso — ISO code for CountryFlag (EU = Schengen)",
        " * @property {string} hint",
        " * @property {{ visaFree?: string[], visaOnArrival?: string[], eVisa?: string[] }} grants",
        " */",
        "",
        "/** @type {HeldVisaOption[]} */",
        f"export const HELD_VISA_OPTIONS = {json.dumps(options, indent=2, ensure_ascii=False)};",
        "",
        """const STATUS_RANK = {
  'visa-free': 0,
  'visa-on-arrival': 1,
  'e-visa': 2,
  'visa-required': 3,
  unknown: 4,
};

/**
 * @param {string} heldVisaId
 * @returns {HeldVisaOption|undefined}
 */
export function getHeldVisaById(heldVisaId) {
  return HELD_VISA_OPTIONS.find((v) => v.id === heldVisaId);
}

/**
 * @param {string} heldVisaId
 * @param {string} destinationIso
 * @returns {'visa-free'|'visa-on-arrival'|'e-visa'|null}
 */
export function lookupHeldVisaGrant(heldVisaId, destinationIso) {
  const visa = getHeldVisaById(heldVisaId);
  if (!visa) return null;
  if (visa.grants.visaFree?.includes(destinationIso)) return 'visa-free';
  if (visa.grants.visaOnArrival?.includes(destinationIso)) return 'visa-on-arrival';
  if (visa.grants.eVisa?.includes(destinationIso)) return 'e-visa';
  return null;
}

/**
 * @param {'visa-free'|'visa-on-arrival'|'e-visa'|'visa-required'|'unknown'} a
 * @param {'visa-free'|'visa-on-arrival'|'e-visa'|'visa-required'|'unknown'} b
 */
export function pickBetterStatus(a, b) {
  return STATUS_RANK[a] <= STATUS_RANK[b] ? a : b;
}
""",
    ]
    OUT_HELD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")

    if not NAME_TO_ISO_PATH.exists():
        raise SystemExit("Run npm run generate:destinations first (missing country_name_to_iso.json)")

    name_to_iso = load_name_to_iso()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    visa_free, voa, evisa = parse_matrix_sheet(wb, name_to_iso)
    costs = parse_costs_sheet(wb, name_to_iso)
    reconcile_in_costs(visa_free, voa, evisa, costs)

    schengen: dict = {"visaFree": ["FR", "ES", "IT", "PT", "GR", "HR", "PL", "DE", "NL", "BE", "AT", "CH", "CZ", "SE", "NO", "DK", "FI", "IE", "LU", "IS", "EE", "LV", "LT", "SK", "SI", "HU", "MT", "CY"]}
    uk: dict = {}
    us: dict = {}
    parse_bonus_sheet(wb, "🇬🇧 UK Visa Bonus", uk, name_to_iso)
    parse_bonus_sheet(wb, "🇺🇸 US Visa Bonus", us, name_to_iso)
    parse_bonus_sheet(wb, "📑 Schengen Bonus", schengen, name_to_iso)

    # Canada / UAE / Australia grants (curated; not in xlsx bonus tabs)
    canada = {"visaFree": ["MX", "CR"], "visaOnArrival": ["CO"]}
    uae = {"visaFree": ["AE", "JO", "MA"], "visaOnArrival": ["KE", "LK"]}
    australia = {"visaFree": ["NZ"], "visaOnArrival": ["ID", "PH", "LK"]}

    write_matrix(visa_free, voa, evisa)
    write_costs(costs)
    write_held(schengen, uk, us, canada, uae, australia)

    print(f"Wrote {OUT_MATRIX.name}")
    print(f"Wrote {OUT_COSTS.name}")
    print(f"Wrote {OUT_HELD.name}")
    print(f"Passports in matrix: {sorted(visa_free.keys())}")
    print(f"Cost entries: {len(costs)}")


if __name__ == "__main__":
    main()
